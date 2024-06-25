#!/usr/bin/env python3

# import sqlite3
import pandas
import random
import numpy as np
import openpyxl
import openpyxl.utils.dataframe
from collections import OrderedDict
from sklearn.ensemble import RandomForestRegressor
from sklearn.inspection import permutation_importance
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib import rcParams
import joblib


# pip install "pandas<2.0.0"
# python -m pip install openpyxl
# ## ## ### important- note that sklearn must be scikit-learn
# python -m pip install scikit-learn
# python -m pip install matplotlib
# python -m pip install seaborn
# python -m pip install joblib


def load_and_format_training_dataset_excel(excelFileName):
    if not (set(['pandas']).issubset(set(globals()))):
        print("Error- modules not loaded. needs 'openpyxl'")
        return()
    TrainingSpeciesAndRatings = pandas.read_excel(excelFileName, sheet_name='TrainingSpeciesAndRatings', index_col='speciesID')
    TrainingSpeciesAndRatings = TrainingSpeciesAndRatings[TrainingSpeciesAndRatings.include != 0]
    TrainingSpeciesAndRatings = TrainingSpeciesAndRatings.drop(columns=['vernacularName', 'distribution','plantHosts','include','notes'])
    Questions = pandas.read_excel(excelFileName, sheet_name='Questions', index_col='questionID')
    Questions = Questions[Questions.include != 0]
    trainingData = pandas.read_excel(excelFileName, sheet_name='Responses', index_col='responseID')
    trainingData = trainingData.drop(columns=['scientificName', 'question','lookupCombo','notes','references'])
    trainingData = trainingData[trainingData['questionID'].isin(list(Questions.index.values))]
    trainingData = trainingData[trainingData['taxonID'].isin(list(TrainingSpeciesAndRatings.index.values))]
    trainingData =  trainingData.pivot_table('answer', 
        index=['taxonID'],
        columns='questionID')
    if len(trainingData[trainingData.isnull().any(axis=1)]) > 0:
        print("WARNING- " + str(len(trainingData[trainingData.isnull().any(axis=1)])) + " species do not have complete data. They are excluded")
        print(trainingData[trainingData.isnull().any(axis=1)])
        trainingData = trainingData.dropna()
    TrainingSpeciesAndRatings['imp2'] = np.where(TrainingSpeciesAndRatings['impactRating'] >= 2, 1, 0)
    TrainingSpeciesAndRatings['imp3'] = np.where(TrainingSpeciesAndRatings['impactRating'] >= 3, 1, 0)
    TrainingSpeciesAndRatings['imp4'] = np.where(TrainingSpeciesAndRatings['impactRating'] >= 4, 1, 0)
    TrainingSpeciesAndRatings['imp5'] = np.where(TrainingSpeciesAndRatings['impactRating'] >= 5, 1, 0)
    TrainingSpeciesAndRatings['imp6'] = np.where(TrainingSpeciesAndRatings['impactRating'] >= 6, 1, 0)
    #reset index of questions to 
    #temporarily reset the index becasue otherwise the index is lost in the merge!
    TrainingSpeciesAndRatings.reset_index(inplace=True)
    dataSummary = TrainingSpeciesAndRatings.merge(trainingData,left_on='speciesID', right_on='taxonID')  
    dataSummary.set_index('speciesID',inplace=True)
    return(dataSummary)

def make_five_random_forest_models(trainingDataSummary):
    #check required modleus are loaded
    if not (set(['RandomForestRegressor', 'pandas', 'np']).issubset(set(globals()))):
        print("Error- modules not loaded. needs 'RandomForestRegressor', 'pandas', 'np'")
        return()
    #make an empty dictionary that will contain the models
    modelDict = dict()
    questionList = list(set(trainingDataSummary.columns) - set(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6']))
    if not (all(isinstance(item, int) for item in questionList)):
        print("Error- dataset has extra columns which are not questionIDs")
        return()   
    trainingData = np.array(trainingDataSummary.drop(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6'], axis=1))
    # note, previously I included the variable random_state = 643765437 for reproducability.)
    for impactLevel in (['imp2','imp3','imp4','imp5','imp6']):
        rf = RandomForestRegressor(n_estimators = 1000)
        labels = np.array(trainingDataSummary[impactLevel])
        rf.fit(trainingData, labels)
        modelDict[impactLevel] = rf
    return(modelDict)

def get_test_data_from_training_dataset(trainingDataSummary,listOfTaxa=[]):
    if not (set(['pandas']).issubset(set(globals()))):
        print("'pandas'")
        return()
    #allows for a single taxonID to be passed.  
    if (isinstance(listOfTaxa,int) or isinstance(listOfTaxa,str)):
        listOfTaxa = [int(listOfTaxa)]
    trainingData = trainingDataSummary.drop(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6'], axis=1)
    if listOfTaxa != []:
        trainingData = trainingData.loc[listOfTaxa]
    return(trainingData)

def make_predictions_from_test_table(testData, modelDict):
    #note, the testData must have the taxon ID as rownames
    taxonIDs = list(testData.index)
    #make up a nested dictionary
    #first level is the impact ratings
    #second level is the different taxa
    predictionsDict = OrderedDict()
    for impactLevel in (['imp2','imp3','imp4','imp5','imp6']):
        predictionsDict[impactLevel] = {}
        tempPredictions = modelDict[impactLevel].predict(testData)
        for i in range(0,len(taxonIDs)):
            predictionsDict[impactLevel][taxonIDs[i]] = tempPredictions[i]
    tempDataFrame = pandas.DataFrame(data=predictionsDict)
    return(tempDataFrame)

def get_means_from_training_dataset(trainingDataSummary):
    df = trainingDataSummary[['imp2', 'imp3', 'imp4', 'imp5', 'imp6']].copy()
    means = df.mean()
    return(means)

def k_equals_n_validation(trainingDataSummary):
    # This will take in the training summary, then take the taxa, and then for
    # each taxon, redevelop the model without it.
    #most of the coputational time is the random forests step.
    # consider changing the input to be only training data summary, but
    # that will involve changing the functions that it calls
    allTrainingSpeciesList = list(trainingDataSummary.index.values)
    allPredictionsDict = {'taxonID':[],'imp2':[],'imp3':[],'imp4':[],'imp5':[],'imp6':[]}
    allMeansDict = {'taxonID':[],'imp2':[],'imp3':[],'imp4':[],'imp5':[],'imp6':[]}
    for taxonID in allTrainingSpeciesList:
        print('developing model for testing taxon:' + str(taxonID))
        testData = get_test_data_from_training_dataset(trainingDataSummary,[taxonID])
        listOfTrainingTaxa = [x for x in allTrainingSpeciesList if x != taxonID]
        trainingData = trainingDataSummary.loc[listOfTrainingTaxa]
        trainingDataMeans = get_means_from_training_dataset(trainingData)
        tempModelsDict = make_five_random_forest_models(trainingDataSummary)
        tempPredictions = make_predictions_from_test_table(testData, tempModelsDict)
        # The following may not be the most efficient way to essentially join 
        # many small data frames together, first converting to a dictionary, 
        # Then looping though the items appending to a dictioray of lists
        #for reading into a data frame again.
        tempPredictionsAsDict = tempPredictions.to_dict(orient='list')
        allPredictionsDict['taxonID'].append(taxonID)
        allMeansDict['taxonID'].append(taxonID)
        for pred in tempPredictionsAsDict:
            allPredictionsDict[pred].append(tempPredictionsAsDict[pred][0])
            allMeansDict[pred].append(trainingDataMeans[pred])
    allPredictions = pandas.DataFrame(data=allPredictionsDict,index=allPredictionsDict['taxonID'])
    allMeans = pandas.DataFrame(data=allMeansDict,index=allMeansDict['taxonID'])
    return(allPredictions,allMeans)

def incremental_validation(trainingDataSummary):
    #this function perfoms the validation tests and is very slow.
    #Currently, it will only retrun a summary, not all of the model runs data, which
    # is very large.
    #
    #trainingDataSummary = load_and_format_training_dataset(sqliteFileName)
    #create a list where the model data will be stored
    modelRunsDict = dict({'repeatNum':[], 'numTestSpecies':[], 'numTrainingSpecies':[],'impactLevel':[], 'method':[], 'tp':[],'fn':[],'fp':[],'tn':[]})
    numRepeats = 10
    #note, numRepeats is approximate, and all are randomly selected, so average would be numRepeats, but will be tested less.
    totalSpecies = len(trainingDataSummary)
    #for numTestSpecies in [2,4,6,8,10,12,14,16,18,20,22,24,26,28,30,32,34,36,38,40]:
    for numTestSpecies in [10,20,30,40,50,60]:
    #for numTestSpecies in [10,20,40]:
        ##for (numTestSpecies in c(1,10,20,30,40)) {
        ##for a range of 1 to numRepeats
        ##numTestSpecies = 10
        #the number of repeats is scaled so that when larger samples are taken, fewer repeats are needed.
        # 1 test = 500 runs (on average sample is a test 10 times)
        # 2 test = 250 runs (on average sample is a test 10 times)
        # 10 tests = 50 runs (on average sample is a test 10 times)
        numTrainingSpecies = totalSpecies - numTestSpecies
        numRepeatesScaled = int((totalSpecies/numTestSpecies)*numRepeats)
        if (numRepeatesScaled == 0):
            numRepeatesScaled = 1
        ##print(numRepeatesScaled)
        # for (repeatNum in 1:numRepeatesScaled) {
        for repeatNum in range(0,numRepeatesScaled):
            #note that the repeatnum is now zero-indexed
            print(str(numTestSpecies) + ":" + "repeat "+ str(repeatNum) + " of " + str(numRepeatesScaled))
            testSample = trainingDataSummary.sample(n=numTestSpecies)
            testSample =  testSample.drop(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6'], axis=1)
            trainingSample = trainingDataSummary.drop(testSample.index)
            tempModels = make_five_random_forest_models(trainingSample)
            tempPredictions = make_predictions_from_test_table(testSample, tempModels)
            tempPredictions.loc[tempPredictions['imp2'] > 0.5, 'imp2_over50'] = 1
            tempPredictions.loc[tempPredictions['imp3'] > 0.5, 'imp3_over50'] = 1
            tempPredictions.loc[tempPredictions['imp4'] > 0.5, 'imp4_over50'] = 1
            tempPredictions.loc[tempPredictions['imp5'] > 0.5, 'imp5_over50'] = 1
            tempPredictions.loc[tempPredictions['imp6'] > 0.5, 'imp6_over50'] = 1
            tempPredictions = tempPredictions.fillna(0)
            for impactRating in ['imp2','imp3','imp4','imp5','imp6']:
                tempPredictions = tempPredictions.astype({impactRating + '_over50': int})
            tempPredictions = tempPredictions.join(trainingDataSummary[['imp2','imp3','imp4','imp5','imp6']].add_suffix('_actual'))
            for impactRating in ['imp2','imp3','imp4','imp5','imp6']:
                RFDT_tp = len(tempPredictions.query(impactRating + '_actual' + '==1 & ' + impactRating + '_over50'+ '==1 '))
                RFDT_fn = len(tempPredictions.query(impactRating + '_actual' + '==0 & ' + impactRating + '_over50'+ '==1 '))
                RFDT_fp = len(tempPredictions.query(impactRating + '_actual' + '==1 & ' + impactRating + '_over50'+ '==0 '))
                RFDT_tn = len(tempPredictions.query(impactRating + '_actual' + '==0 & ' + impactRating + '_over50'+ '==0 '))
                modelRunsDict['repeatNum'].append(repeatNum)
                modelRunsDict['numTestSpecies'].append(numTestSpecies)
                modelRunsDict['numTrainingSpecies'].append(numTrainingSpecies)
                modelRunsDict['impactLevel'].append(impactRating)
                modelRunsDict['method'].append('random decision tree forest')
                modelRunsDict['tp'].append(RFDT_tp)
                modelRunsDict['fn'].append(RFDT_fn)
                modelRunsDict['fp'].append(RFDT_fp)
                modelRunsDict['tn'].append(RFDT_tn)
    newDataFrame = pandas.DataFrame(data=modelRunsDict)
    temp = newDataFrame.groupby(['numTrainingSpecies','method','impactLevel'])
    # temp.first()
    # temp.sum()
    summaryAsDict = temp.agg(sum).reset_index().to_dict(orient='list')
    numRows = len(summaryAsDict['method'])
    #create an empy list for all items
    summaryAsDict['numCorrect'] = [None] * numRows
    summaryAsDict['accuracy'] = [None] * numRows
    summaryAsDict['specificity'] =  [None] * numRows
    summaryAsDict['sensitivity'] =  [None] * numRows
    summaryAsDict['falseNegativeRate'] =  [None] * numRows
    summaryAsDict['MCC'] =  [None] * numRows
    summaryAsDict['totalTested'] =  [None] * numRows
    for i in range(0,numRows):
        TP = summaryAsDict['tp'][i]
        FN = summaryAsDict['fn'][i]
        FP = summaryAsDict['fp'][i]
        TN = summaryAsDict['tn'][i]
        #print(str(TP) + " " + str(FN))
        # numberTested = sum([summaryAsDict['tp'])
        summaryAsDict['totalTested'][i] = (TP+FN+FP+TN)
        summaryAsDict['numCorrect'][i] = (TP + TN)
        summaryAsDict['accuracy'][i] = ((TP + TN) / (TP + TN + FP + FN ))
        summaryAsDict['specificity'][i] = (TN / (TN + FP))
        summaryAsDict['sensitivity'][i] = (TP / (TP + FN))
        summaryAsDict['falseNegativeRate'][i] = (FN / (FN + TP))
        summaryAsDict['MCC'][i] = (((TP * TN)-(FP * FN))/ np.sqrt((TP + FP)*(TP + FN)*(TN + FP)*(TN + FN)))
    #ModelRunsSummaryDict= dict({'method':[], 'numTestSpecies':[], 'numTrainingSpecies':[],'impactLevel':[], 'totalTested':[],'numCorrect':[],'accuracy':[], 'specificity':[], 'sensitivity':[], 'falseNegativeRate':[], 'MCC':[], 'tp':[],'fn':[],'fp':[],'tn':[]})
    del summaryAsDict['repeatNum']
    del summaryAsDict['numTestSpecies']
    modelRunsSummary = pandas.DataFrame(data=summaryAsDict)
    return(newDataFrame,modelRunsSummary)


def get_importance_of_questions_all_taxa(trainingDataSummary,modelDict,questionLookupDict):
    # make empty dictionary
    questionList = sorted(list(set(trainingDataSummary.columns) - set(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6'])))
    questionValueDict = dict({'questionID':[], 'impactRating':[], 'importances_mean':[], 'importances_std':[], 'questionShort':[]})
    answerMatirx = trainingDataSummary.drop(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6'], axis=1)
    for impactRating in ['imp2','imp3','imp4','imp5','imp6']:
        #perfrom the permutation_importance for the predictions at that impact level
        permImportance = permutation_importance(modelDict[impactRating], answerMatirx, trainingDataSummary[impactRating].tolist(), n_repeats=10, random_state=86538)
        #simply append all of the results lists, as well as the question Ids and 
        #impact ratings to the dictionary
        questionValueDict['questionID'].extend(questionList)
        questionValueDict['impactRating'].extend([impactRating]* len(questionList))
        questionValueDict['importances_mean'].extend(permImportance['importances_mean'])
        questionValueDict['importances_std'].extend(permImportance['importances_std'])
        for questionID in questionList:
            questionValueDict['questionShort'].append(questionLookupDict[questionID])
        #questionValueDict['importances'].extend(permImportance['importances'])
        questionValueDataFrame = pandas.DataFrame(data=questionValueDict)
    return(questionValueDataFrame)

def load_questions_dict(excelFileName):
    if not (set(['pandas']).issubset(set(globals()))):
        print("Error- modules not loaded. needs 'openpyxl'")
        return()
    questionDataFrame = pandas.read_excel(excelFileName, sheet_name='Questions', index_col='questionID')
    questionDataFrame = questionDataFrame[questionDataFrame.include != 0]
    questionDataFrame = questionDataFrame.drop(columns=['include'])
    #questionDict = dict(zip(questionDataFrame['questionID'], questionDataFrame['questionShort']))
    #note, to_dict makes the colums the index and the column values tthe keys. the T below transposes the data frame so that the columns are questions
    #questionDict = questionDataFrame.T.to_dict()
    questionDict = dict(questionDataFrame.reset_index().values)
    return(questionDict)

def make_five_random_forest_models(trainingDataSummary):
    #check required modleus are loaded
    if not (set(['RandomForestRegressor', 'pandas', 'np']).issubset(set(globals()))):
        print("Error- modules not loaded. needs 'RandomForestRegressor', 'pandas', 'np'")
        return()
    #make an empty dictionary that will contain the models
    modelDict = dict()
    questionList = list(set(trainingDataSummary.columns) - set(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6']))
    if not (all(isinstance(item, int) for item in questionList)):
        print("Error- dataset has extra columns which are not questionIDs")
        return()   
    trainingData = np.array(trainingDataSummary.drop(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6'], axis=1))
    # note, previously I included the variable random_state = 643765437 for reproducability.)
    for impactLevel in (['imp2','imp3','imp4','imp5','imp6']):
        rf = RandomForestRegressor(n_estimators = 1000)
        labels = np.array(trainingDataSummary[impactLevel])
        rf.fit(trainingData, labels)
        modelDict[impactLevel] = rf
    return(modelDict)
 
def get_mean_values_by_question(trainingDataSummary):
    #check required modleus are loaded
    if not (set(['pandas', 'np']).issubset(set(globals()))):
        print("Error- modules not loaded. needs 'pandas', 'np'")
        return()
    #make an empty dictionary that will contain the models
    questionList = list(set(trainingDataSummary.columns) - set(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6']))
    if not (all(isinstance(item, int) for item in questionList)):
        print("Error- dataset has extra columns which are not questionIDs")
        return()
    trainingData = trainingDataSummary.drop(['scientificName','impactRating','imp2','imp3','imp4','imp5','imp6'], axis=1)
    means = trainingData.mean()
    means = means.to_dict()
    return(means)

def generate_working_sheet_openpyxl(questionsDict,trainingDataSummary,randomForestModelsDict,meansDict,outputFileName):
    from datetime import datetime
    import openpyxl
    #import openpyxl.utils.dataframe
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule, ColorScaleRule, ColorScale, FormatObject
    from openpyxl.styles import Alignment, Color, PatternFill, Font, Border
    from sklearn.ensemble import RandomForestRegressor
    from sklearn.tree import _tree
    numRepeats = 1000
    
    redFill = PatternFill(start_color='EE1111',
                    end_color='EE1111',
                    fill_type='solid')
    
    def tree_to_excel_code(tree, cellCoods):
        '''
        Output: an excel formula for the one tree
        Input tree: decision tree model
        Input cellCoods: list the excel-formatted coodinates in the order that the model is trained(e.g. 'B2','C2','D2'....)
        '''
        tree_ = tree.tree_
        #it takes the question list, 
        # and looks at some feature value, which I sewe as -2, but
        # the _tree.TREE_UNDEFINED seems to see as False
        # the quesionnums are then repalced with undefined
        # this indicates where to terminate the if statement
        feature_name = []
        for fea in tree_.feature:
            if fea != _tree.TREE_UNDEFINED:
                feature_name.append(cellCoods[fea])
            else:
                feature_name.append("undefined!")
        excelFormulaStr = '='
        def recurse(node):
            #indent = "  " * depth
            tempStr = ''
            if tree_.feature[node] != _tree.TREE_UNDEFINED:
                name = feature_name[node]
                # since we only have 0 or 1, we dont need a theshold
                #threshold = tree_.threshold[node]
                tempStr += "if({},".format(name)
                tempStr += recurse(tree_.children_right[node])
                tempStr += ","
                tempStr += recurse(tree_.children_left[node])
                tempStr += ")"
            else:
                tempStr += "{}".format(int(tree_.value[node][0][0]))
            return(tempStr)
        #print("def tree({}):".format(", ".join(feature_names)))
        excelFormulaStr += recurse(0)
        return(excelFormulaStr)
    wb = openpyxl.Workbook()
    #note, the line below edits the file, but doesnt seem to work- not sure what is wrong [this may work now, check]
    wb.calculation.calcMode = "manual"
    # Worksheet 0: Instructions. #####################################################
    ws0 = wb.create_sheet()
    ws0.title = "Instructions"
    ws0.column_dimensions['A'].width = 160
    ws0.append(["Woodborer-specific risk assessment model."])
    ws0['A' + str(ws0.max_row)].font = Font(size=20)
    ws0.append(["Version: " + datetime.today().strftime('%Y%m%d')])
    ws0.append([""])
    ws0.append(["Important note: The default behavior of Microsoft excel is to apply a single calculation option to all open workbooks. So although this sheet is set to calculate manually (by clicking -calculate now- in the formula tab), if you open at the same time as another workbook, the calculation will be automatic, recalculating 1000000s of values at every change, which will be very slow. The work around for this is to do one of the following: a) close other workbooks, b) set other workbooks to calculate manually, or c) find out how to open a separate instance of Excel"])
    ws0.append([""])
    ws0.append(["Instructions"])
    ws0['A' + str(ws0.max_row)].font = Font(size=20)
    ws0.append(["Background data sheet: First fill out the background data, which is information about the current distribution, host and affected commodities. This has no influence on the model but may be useful when interpreting the model for high impact pests. Entering a taxon name (species or even variety) is essential, the other fields are optional but recommended."])
    ws0.append(["Input sheet: The predictive variables are questions which you can answer true or false. The answer is a number which is the probability of it being true. If you do not know, leave the number blank- it will use the average from the species already introduced in north America as its prior probability"])
    ws0.append([""])
    ws0.append(["""As a guide to describing certainty values, the following terms can be used as an aid when formulating the probability:
    1:            Certainly true
    0.99 to 1:    Extremely likely to be true
    0.9 to 0.99:  Very likely to be true
    0.66 to 0.9:  Likely to be true
    0.33 to 0.66: About as likely true as false
    0.1 to 0.33:  likely to be false
    0.01 to 0.1:  Very likely to be false
    0 to 0.01:    Extremely likely to be false
    0:            Certainly false
    You must choose an arbitrary value for the probability which matches your interpretation of the literature.
    If the value is truly unknown, and you have no information leave the answer blank. The simulation will use a default value taken from existing species already introduced.
    """])
    ws0.append([""])
    ws0.append(["""Once you have completed filling in the worksheet, under the "Formulas" tab on MS excel, click "calculate now".

    The high number of simulations means that there should be minimal variation between runs. However, the user must not re-run the simulation multiple times with the intent of choosing a particular result.

    View the results in the "Report" sheet
    """])
    for row in ws0.iter_rows():
        for cell in row:      
            cell.alignment = Alignment(wrapText=True)
    ws0.protection.sheet = True
    # Worksheet 1: Background. #####################################################
    ws1 = wb.active
    ws1.title = "Background"
    backgroundData = [
        ['Scientific Name', ''],
        ['Authority', ''],
        ['Current distribution', ''],
        ['Host Range', ''],
        ['Notes', ''],
        ['taxonID', ''],
    ]
    # add column headings
    ws1.append(["Field Name", "input"])
    for row in backgroundData:
        ws1.append(row)
    tab1 = openpyxl.worksheet.table.Table(displayName="BackgroundData", ref="A1:B7")
    # Add a default style with striped rows and banded columns
    style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab1.tableStyleInfo = style
    #Table must be added using ws.add_table() method to avoid duplicate names.
    #Using this method ensures table name is unque through out defined names and all other table name. 
    ws1.add_table(tab1)
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 50
    ws1.protection.sheet = True
    for x in range(2, ws1.max_row+1):
        y = 2
        ws1.cell(row = x, column = y).protection = openpyxl.styles.Protection(locked=False, hidden=False)
    # Worksheet 2: questions and answers############################################
    ws2 = wb.create_sheet()
    ws2.title = "Input"
    # columsn will be questionI
    ws2.append(["questionID", "questionShort", "entryAnswer", "manualReference", "notes", "answerWithUncertainty", "defaultAnswer","uncertainty"])
    for i, questionID in enumerate(questionsDict):
        row = [questionID, questionsDict[questionID], None, None, None, '=IF(C{0}="",G{0},C{0})'.format(str(i+2)),meansDict[questionID], '=(0.5-ABS(F{0} - 0.5))*2'.format(str(i+2))]
        ws2.append(row)
    tableRefs = ("A1:H" + str(len(questionsDict) + 1))
    tab2 = openpyxl.worksheet.table.Table(displayName="InputData", ref=tableRefs)
    ws2.add_table(tab2)
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 40
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 15
    ws2.column_dimensions['H'].width = 15
    #wrap ext in the question
    for row in ws2[2:ws2.max_row + 1]:  # skip the header
        cell = row[1]             # column H
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    tab2.tableStyleInfo = style
    #protect all except input variables
    ws2.protection.sheet = True
    for x in range(2, ws2.max_row):
        for y in (3,4,5):
            ws2.cell(row = x, column = y).protection = openpyxl.styles.Protection(locked=False, hidden=False)
    # Format using a formula highlighting invalid entries
    inputCellRange = "C2:C" + str(ws2.max_row)
    ws2.conditional_formatting.add(inputCellRange, FormulaRule(formula=['OR(C2<0,C2>1)'], stopIfTrue=True, fill=redFill))
    #hide columns users do not need to see
    for col in ['A', 'F', 'G', 'H']:
        ws2.column_dimensions[col].hidden= True
    # Worksheet 3: Simulations #####################################################
    #ws3 is the simulation data, columns are questions, rows are simulations
    ws3 = wb.create_sheet()
    ws3.title = "Sim"
    simulationRowAsList = []
    for i, questionID in enumerate(questionsDict):
        #note, first column will be the repeat number, so all indexes will need to be shifted
        #note2, the location of the cell containing data will be Input!{0}{1}".format("F",str(i+2)
        simulationRowAsList.append("=IF(RAND()>Input!{0}{1},0,1)".format("F",str(i+2)))
    listOfQuestions = questionsDict.keys()
    questionStrList = ["q" + str(i) for i in listOfQuestions]
    ws3.append(["repeat"] + questionStrList)
    for i in range(1,numRepeats+1):
        ws3.append([i] + simulationRowAsList)
    ws3.protection.sheet = True
    ws3.sheet_state = 'hidden'
    # Worksheet 4: Random forest for Imp2 ##########################################
    #ws4 is the random forest for imp2
    #tree_to_code(tree,questionListStr)
    # tree_to_excel_code(tree, questionListStr)
    # tree_to_code(tree,questionListStr)   
    #note, the below can be made much more efficeintly since it re-calculated the
    # excel formula for eveyr cell, despo=ite being consistem among columns
    ws4 = wb.create_sheet()
    ws4.title = "RDTF_imp2"
    randomForestModel = randomForestModelsDict["imp2"]
    rowToAdd = ["repeat"]
    for i in range (0,len(randomForestModel.estimators_)):
        rowToAdd.append("tree" + str(i+1).zfill(4))
    ws4.append(rowToAdd)
    for repeatNum in range(1,numRepeats+1):
        #make the list of coodinates corresponding to the list of responses for that simulation
        rowToAdd = [repeatNum]
        tempCoodList = []
        for i in range(0,len(listOfQuestions)):
            tempCoodList.append("Sim!" + get_column_letter(i+2) + str(repeatNum + 1))      
        for i, estimator in enumerate(randomForestModel.estimators_):
            formulaToInsert = tree_to_excel_code(estimator,tempCoodList)
            rowToAdd.append(formulaToInsert)
        #print(rowToAdd)
        ws4.append(rowToAdd)  
    ws4.protection.sheet = True
    ws4.sheet_state = 'hidden'
    ws5 = wb.create_sheet()
    ws5.title = "RDTF_imp3"
    randomForestModel = randomForestModelsDict["imp3"]
    rowToAdd = ["repeat"]
    for i in range (0,len(randomForestModel.estimators_)):
        rowToAdd.append("tree" + str(i+1).zfill(4))
    ws5.append(rowToAdd)
    for repeatNum in range(1,numRepeats+1):
        #make the list of coodinates corresponding to the list of responses for that simulation
        rowToAdd = [repeatNum]
        tempCoodList = []
        for i in range(0,len(listOfQuestions)):
            tempCoodList.append("Sim!" + get_column_letter(i+2) + str(repeatNum + 1))      
        for i, estimator in enumerate(randomForestModel.estimators_):
            formulaToInsert = tree_to_excel_code(estimator,tempCoodList)
            rowToAdd.append(formulaToInsert)
        #print(rowToAdd)
        ws5.append(rowToAdd)
    ws5.protection.sheet = True
    ws5.sheet_state = 'hidden'
    ws6 = wb.create_sheet()
    ws6.title = "RDTF_imp4"
    randomForestModel = randomForestModelsDict["imp4"]
    rowToAdd = ["repeat"]
    for i in range (0,len(randomForestModel.estimators_)):
        rowToAdd.append("tree" + str(i+1).zfill(4))
    ws6.append(rowToAdd)
    for repeatNum in range(1,numRepeats+1):
        #make the list of coodinates corresponding to the list of responses for that simulation
        rowToAdd = [repeatNum]
        tempCoodList = []
        for i in range(0,len(listOfQuestions)):
            tempCoodList.append("Sim!" + get_column_letter(i+2) + str(repeatNum + 1))      
        for i, estimator in enumerate(randomForestModel.estimators_):
            formulaToInsert = tree_to_excel_code(estimator,tempCoodList)
            rowToAdd.append(formulaToInsert)
        #print(rowToAdd)
        ws6.append(rowToAdd)
    ws6.protection.sheet = True
    ws6.sheet_state = 'hidden'
    ws7 = wb.create_sheet()
    ws7.title = "RDTF_imp5"
    randomForestModel = randomForestModelsDict["imp5"]
    rowToAdd = ["repeat"]
    for i in range (0,len(randomForestModel.estimators_)):
        rowToAdd.append("tree" + str(i+1).zfill(4))
    ws7.append(rowToAdd)
    for repeatNum in range(1,numRepeats+1):
        #make the list of coodinates corresponding to the list of responses for that simulation
        rowToAdd = [repeatNum]
        tempCoodList = []
        for i in range(0,len(listOfQuestions)):
            tempCoodList.append("Sim!" + get_column_letter(i+2) + str(repeatNum + 1))      
        for i, estimator in enumerate(randomForestModel.estimators_):
            formulaToInsert = tree_to_excel_code(estimator,tempCoodList)
            rowToAdd.append(formulaToInsert)
        #print(rowToAdd)
        ws7.append(rowToAdd)
    ws7.protection.sheet = True
    ws7.sheet_state = 'hidden'
    ws8 = wb.create_sheet()
    ws8.title = "RDTF_imp6"
    randomForestModel = randomForestModelsDict["imp6"]
    rowToAdd = ["repeat"]
    for i in range (0,len(randomForestModel.estimators_)):
        rowToAdd.append("tree" + str(i+1).zfill(4))
    ws8.append(rowToAdd)
    for repeatNum in range(1,numRepeats+1):
        #make the list of coodinates corresponding to the list of responses for that simulation
        rowToAdd = [repeatNum]
        tempCoodList = []
        for i in range(0,len(listOfQuestions)):
            tempCoodList.append("Sim!" + get_column_letter(i+2) + str(repeatNum + 1))      
        for i, estimator in enumerate(randomForestModel.estimators_):
            formulaToInsert = tree_to_excel_code(estimator,tempCoodList)
            rowToAdd.append(formulaToInsert)
        #print(rowToAdd)
        ws8.append(rowToAdd)
    ws8.protection.sheet = True
    ws8.sheet_state = 'hidden'
    ws9 = wb.create_sheet()
    ws9.title = "totals"
    ws9.append(["repeat","imp2","imp3","imp4","imp5","imp6"])
    firstColumnLetter =  get_column_letter(2)
    lastColumnLetter =  get_column_letter(len(randomForestModel.estimators_) + 1)
    for repeatNum in range(1,numRepeats+1):
        rowToAdd = [repeatNum]
        for impactRating in ["imp2","imp3","imp4","imp5","imp6"]:
            formulaToInsert = "=AVERAGE(RDTF_{0}!{1}{2}:{3}{2})".format(impactRating,firstColumnLetter,str(repeatNum+1),lastColumnLetter)
            rowToAdd.append(formulaToInsert)
        ws9.append(rowToAdd)
    ws9.protection.sheet = True
    ws9.sheet_state = 'hidden'
    ws10 = wb.create_sheet()
    ws10.title = "simulationSummary"
    ws10.append(["","imp2","imp3","imp4","imp5","imp6"])
    rowToAdd = ["Mean"]
    for i, impactRating in enumerate(["imp2","imp3","imp4","imp5","imp6"]):
        columnLetter = get_column_letter(i+2)
        firstCell = columnLetter + str(2)
        lastCell = columnLetter  + str(numRepeats+1)
        formulaToInsert = "=AVERAGE({0}!{1}:{2})".format("totals",firstCell,lastCell)
        rowToAdd.append(formulaToInsert)
    ws10.append(rowToAdd)
    rowToAdd = []
    for percentile in [0.0,0.05,0.25,0.5,0.75,0.95,1.0]:
        rowToAdd = [str(percentile)]
        for i, impactRating in enumerate(["imp2","imp3","imp4","imp5","imp6"]):
            columnLetter = get_column_letter(i+2)
            firstCell = columnLetter + str(1)
            lastCell = columnLetter  + str(numRepeats)
            formulaToInsert = "=PERCENTILE({0}!{1}:{2},{3})".format("totals",firstCell,lastCell,str(percentile))
            rowToAdd.append(formulaToInsert)
        ws10.append(rowToAdd)
    ws10.protection.sheet = True
    #histogram calculations
    ws11 = wb.create_sheet()
    ws11.title = "histogramBins"
    numberOfBins = 40
    ws11.append(["binMax","imp2","imp3","imp4","imp5","imp6"])
    for b in range(0,numberOfBins):
        #note, binMin may be misleading since values must be greater but not equal
        #to binMin, but can be equal to bin Max   
        # note, the round is to avoid floating point number issues
        binMin = round((1/numberOfBins) * (b),4)
        binMax = round((1/numberOfBins) * (b + 1),4)
        #print(str(binMin) + " < count <= " + str(binMax))
        rowToAdd = [str(binMax)]
        for i, impactRating in enumerate(["imp2","imp3","imp4","imp5","imp6"]):
            columnLetter = get_column_letter(i+2)
            firstCell = columnLetter + str(1)
            lastCell = columnLetter  + str(numRepeats)
            #formulaToInsert = "=PERCENTILE({0}!{1}:{2},{3})".format("totals",firstCell,lastCell,str(percentile))
            if b == 0:
                formulaToInsert = "=COUNTIFS({0}!{1}:{2},\"<={3}\")".format("totals",firstCell,lastCell,str(binMax))
            else:
                formulaToInsert = "=COUNTIFS({0}!{1}:{2},\">{3}\",{0}!{1}:{2},\"<={4}\")".format("totals",firstCell,lastCell,str(binMin),str(binMax))    
            rowToAdd.append(formulaToInsert)
        ws11.append(rowToAdd)
    ws11.protection.sheet = True
    ws11.sheet_state = 'hidden'
    ################################################################################
    ws12 = wb.create_sheet()
    ws12.title = "densityBins"
    numberOfBins = 10
    rowToAdd = ["impactRating"]
    for b in range(0,numberOfBins):
        binMax = round((1/numberOfBins) * (b + 1),4)
        rowToAdd.append(binMax)
    ws12.append(rowToAdd)
    for i, impactRating in enumerate(["imp2","imp3","imp4","imp5","imp6"]):
        rowToAdd = [impactRating]
        for b in range(0,numberOfBins):
            binMin = round((1/numberOfBins) * (b),4)
            binMax = round((1/numberOfBins) * (b + 1),4)
            columnLetter = get_column_letter(i+2)
            firstCell = columnLetter + str(1)
            lastCell = columnLetter  + str(numRepeats)
            #formulaToInsert = "=PERCENTILE({0}!{1}:{2},{3})".format("totals",firstCell,lastCell,str(percentile))
            if b == 0:
                formulaToInsert = "=COUNTIFS({0}!{1}:{2},\"<={3}\")/{4}".format("totals",firstCell,lastCell,str(binMax),numRepeats)
            else:
                formulaToInsert = "=COUNTIFS({0}!{1}:{2},\">{3}\",{0}!{1}:{2},\"<={4}\")/{5}".format("totals",firstCell,lastCell,str(binMin),str(binMax),numRepeats) 
            rowToAdd.append(formulaToInsert)
        ws12.append(rowToAdd)
    ws12.protection.sheet = True
    ws12.sheet_state = 'hidden'
    ################################################################################
    ws13 = wb.create_sheet()
    ws13.title = "Report"
    ws13.append(["Results Summary"])
    ws13['A' + str(ws13.max_row)].font = Font(size=20)
    ws13.append(["Potential pest investigated:",'=IF(Background!B2<>"",Background!B2,"")'])
    ws13.append(["Model version: ",datetime.today().strftime('%Y%m%d')])
    ws13.append(["Known distribution: ",'=IF(Background!B4<>"",Background!B4,"none provided")'])
    ws13.append(["Known distribution: ",'=IF(Background!B5<>"",Background!B5,"none provided")'])
    temp = '=({1}-COUNTBLANK(Input!C2:C{0})) & " out of {1} questions answered (= " & round((({1}-COUNTBLANK(Input!C2:C{0})) / {1}),2) * 100 & "%)" '.format(str(len(questionsDict) + 1),str(len(questionsDict)))
    ws13.append(["Information available: ",temp])
    ws13.append(["Average uncertainty: ",'=round((AVERAGE(Input!H2:H{0})) * 100,2) & "%"'.format(str(len(questionsDict) + 1))])
    ws13.append(["Risk of medium impact (3+): ","=round(simulationSummary!C2,3)*100"])
    ws13.append(["Risk of high impact (5+): ","=round(simulationSummary!E2,3)*100"])
    binsHeadingRow = ['Predicted probability (%)']
    for b in range(0,numberOfBins):
        binMax = round((1/numberOfBins) * (b + 1),1)
        binsHeadingRow.append((binMax*100))
    ws13.append(["Simulation results: "])
    ws13['A' + str(ws13.max_row)].font = Font(size=20)
    ws13.append(["Proportion of simulations that predict the probability that the impact matches at least the following statement"])
    ws13.append([])
    ws13.append(["At least impact rating 2: Minor damage; examples: leaf/needle loss, leaf/needle discoloration, twig dieback, or fruit drop."])
    ws13.append(binsHeadingRow)
    rowToAdd = ["Proportion of simulations (%)"]
    i = 0
    for b in range(0,numberOfBins):
        rowToAdd.append("=DensityBins!{0}{1}*100".format(get_column_letter(b+2),str(i+2)))
    ws13.append(rowToAdd)
    rangeForContitionalFormatting= "B{0}:K{0}".format(ws13.max_row)
    ws13.conditional_formatting.add(rangeForContitionalFormatting, ColorScaleRule(start_type='num', start_value=0, start_color='00FFFFFF',
            end_type='num', end_value=50, end_color='008b8680'))
    ws13.append(["Avegarge prediction (%):","=round(simulationSummary!B2,3)*100"])
    ws13['A' + str(ws13.max_row)].font = Font(bold=True)
    ws13['B' + str(ws13.max_row)].font = Font(bold=True)
    ws13.append([])
    ws13.append(["At least impact rating 3: Mortality of individual stressed plants."])
    ws13.append(binsHeadingRow)
    rowToAdd = ["Proportion of simulations (%)"]
    i = 1
    for b in range(0,numberOfBins):
        rowToAdd.append("=DensityBins!{0}{1}*100".format(get_column_letter(b+2),str(i+2)))
    ws13.append(rowToAdd)
    rangeForContitionalFormatting= "B{0}:K{0}".format(ws13.max_row)
    ws13.conditional_formatting.add(rangeForContitionalFormatting, ColorScaleRule(start_type='num', start_value=0, start_color='00FFFFFF',
            end_type='num', end_value=50, end_color='008b8680'))
    ws13.append(["Avegarge prediction (%):","=round(simulationSummary!C2,3)*100"])
    ws13['A' + str(ws13.max_row)].font = Font(bold=True)
    ws13['B' + str(ws13.max_row)].font = Font(bold=True)
    ws13.append([])
    ws13.append(["At least impact rating 4: Weakening of an individual plant that suffers mortality from another agent."])
    ws13.append(binsHeadingRow)
    rowToAdd = ["Proportion of simulations (%)"]
    i = 2
    for b in range(0,numberOfBins):
        rowToAdd.append("=DensityBins!{0}{1}*100".format(get_column_letter(b+2),str(i+2)))
    ws13.append(rowToAdd)
    rangeForContitionalFormatting= "B{0}:K{0}".format(ws13.max_row)
    ws13.conditional_formatting.add(rangeForContitionalFormatting, ColorScaleRule(start_type='num', start_value=0, start_color='00FFFFFF',
            end_type='num', end_value=50, end_color='008b8680'))
    ws13.append(["Avegarge prediction (%):","=round(simulationSummary!D2,3)*100"])
    ws13['A' + str(ws13.max_row)].font = Font(bold=True)
    ws13['B' + str(ws13.max_row)].font = Font(bold=True)
    ws13.append([])
    ws13.append(["At least impact rating 5: Mortality of individual healthy plants."])
    ws13.append(binsHeadingRow)
    rowToAdd = ["Proportion of simulations (%)"]
    i = 3
    for b in range(0,numberOfBins):
        rowToAdd.append("=DensityBins!{0}{1}*100".format(get_column_letter(b+2),str(i+2)))
    ws13.append(rowToAdd)
    rangeForContitionalFormatting= "B{0}:K{0}".format(ws13.max_row)
    ws13.conditional_formatting.add(rangeForContitionalFormatting, ColorScaleRule(start_type='num', start_value=0, start_color='00FFFFFF',
            end_type='num', end_value=50, end_color='008b8680'))
    ws13.append(["Avegarge prediction (%):","=round(simulationSummary!E2,3)*100"])
    ws13['A' + str(ws13.max_row)].font = Font(bold=True)
    ws13['B' + str(ws13.max_row)].font = Font(bold=True)
    ws13.append([])
    ws13.append(["At least impact rating 6: =Isolated or sporadic mortality within an affected plant populationa; examples: occasional outbreaks that yield > 10% mortality, 90% mortality with regeneration, or sustained mortality of 5% per year in multiple populations."])
    ws13.append(binsHeadingRow)
    rowToAdd = ["Proportion of simulations (%)"]
    i = 4
    for b in range(0,numberOfBins):
        rowToAdd.append("=DensityBins!{0}{1}*100".format(get_column_letter(b+2),str(i+2)))
    ws13.append(rowToAdd)
    rangeForContitionalFormatting= "B{0}:K{0}".format(ws13.max_row)
    ws13.conditional_formatting.add(rangeForContitionalFormatting, ColorScaleRule(start_type='num', start_value=0, start_color='00FFFFFF',
            end_type='num', end_value=50, end_color='008b8680'))
    ws13.append(["Avegarge prediction (%):","=round(simulationSummary!F2,3)*100"])
    ws13['A' + str(ws13.max_row)].font = Font(bold=True)
    ws13['B' + str(ws13.max_row)].font = Font(bold=True)
    ws13.column_dimensions['A'].width = 32
    ws13.column_dimensions['B'].width = 5
    ws13.column_dimensions['C'].width = 5
    ws13.column_dimensions['D'].width = 5
    ws13.column_dimensions['E'].width = 5
    ws13.column_dimensions['F'].width = 5
    ws13.column_dimensions['G'].width = 5
    ws13.column_dimensions['H'].width = 5
    ws13.column_dimensions['I'].width = 5
    ws13.column_dimensions['J'].width = 5
    ws13.column_dimensions['K'].width = 5
    temp = """
    Average uncertainty of questions (average distance from 0 or 1)
    Mean probability of being a high impact pest (rating 5 or more):
    Mean probability of being at least a medium impact pest (rating 3 or more)
    Predictive model results:
    Probability of causing at least the following imapcts:
    Impact rating 2: vdsgdsgdsgds
    [divided by percentile, then bin density (proportion of total) is given- essentially a tabular histogram]
    Entry data summary

    5=Mortality of individual healthy plants.
    6=Isolated or sporadic mortality within an affected plant populationa; examples: occasional outbreaks that yield > 10% mortality, 90% mortality with regeneration, or sustained mortality of 5% per year in multiple populations.
    7=Extensive or persistent mortality within a population; example: more than 25% mortality over 10 years.
    8=Wave of plant mortality with regional spread of the insect.
    9=Functional extinction of the host plant.
    """
    wb.save(outputFileName)


def plot_figure_2(dataSummary):
    # Adjust figure size
    plt.figure(figsize=(12, 14))  # Width, Height in inches
    bins = np.arange(0.0, 1.05, 0.05)
    plt.subplot(4, 1, 1)
    dataSummary = dataSummary.assign(Actual_values=dataSummary.imp2.map({0: "No", 1: "Yes"}))
    plt.xlim(0, 1)
    imp2plot = sns.histplot(data=dataSummary, x="RDTF_imp2", hue="Actual_values", fill=True, common_norm=True, alpha=1, multiple="stack",  bins=bins)
    #imp2plot.set(xlabel="Model predictions")
    imp2plot.set(xlabel=None)
    imp2plot.set(ylabel="Count")
    imp2plot.annotate('At least impact rating 2', xy=(0.5, 0.90), xycoords='axes fraction', fontsize=12, ha="center")
    imp2plot.axvline(x=0.5,ymin=0,ymax=100, color='gray', linestyle='--')
    imp2plot.legend(loc='upper right', bbox_to_anchor=(0.95, 1.0), title="Actual values")
    # imp2plot
    plt.subplot(4, 2, 3)
    plt.xlim(0, 1)
    imp3plot = sns.histplot(data=dataSummary, x="RDTF_imp3", hue="imp3", fill=True, common_norm=True, alpha=1, multiple="stack",  bins=bins, legend=False)
    imp3plot.set(xlabel=None)
    imp3plot.set(ylabel=None)
    imp3plot.axvline(x=0.5,ymin=0,ymax=100, color='gray', linestyle='--')
    imp3plot.annotate('At least impact rating 3', xy=(0.5, 0.90), xycoords='axes fraction', fontsize=12, ha="center")
    # imp3plot 
    plt.subplot(4, 2, 4)
    plt.xlim(0, 1)
    imp4plot = sns.histplot(data=dataSummary, x="RDTF_imp4", hue="imp4", fill=True, common_norm=True, alpha=1, multiple="stack", bins=bins, legend=False)
    imp4plot.set(xlabel=None)
    imp4plot.set(ylabel=None)
    imp4plot.axvline(x=0.5,ymin=0,ymax=100, color='gray', linestyle='--')
    imp4plot.annotate('At least impact rating 4', xy=(0.5, 0.90), xycoords='axes fraction', fontsize=12, ha="center")
    # imp4plot
    plt.subplot(4, 2, 5)
    plt.xlim(0, 1)
    imp5plot = sns.histplot(data=dataSummary, x="RDTF_imp5", hue="imp5", fill=True, common_norm=True, alpha=1, multiple="stack", bins=bins, legend=False)
    imp5plot.set(xlabel=None)
    imp5plot.set(ylabel=None)
    imp5plot.axvline(x=0.5,ymin=0,ymax=100, color='gray', linestyle='--')
    imp5plot.annotate('At least impact rating 5', xy=(0.5, 0.90), xycoords='axes fraction', fontsize=12, ha="center")
    # imp5plot
    plt.subplot(4, 2, 6)
    plt.xlim(0, 1)
    imp6plot = sns.histplot(data=dataSummary, x="RDTF_imp6", hue="imp6", fill=True, common_norm=True, alpha=1, multiple="stack", bins=bins, legend=False)
    imp6plot.set(xlabel=None)
    imp6plot.set(ylabel=None)
    imp6plot.axvline(x=0.5,ymin=0,ymax=100, color='gray', linestyle='--')
    imp6plot.annotate('At least impact rating 6', xy=(0.5, 0.90), xycoords='axes fraction', fontsize=12, ha="center")
    # imp6plot
    plt.savefig('testingFigure.pdf')

